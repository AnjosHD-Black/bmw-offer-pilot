from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.worksheet.pagebreak import Break
from datetime import datetime
import os


def underline(ws, cell):
    ws[cell].font = Font(underline="single", bold=True)


def bottom_border(ws, row):
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws[f"{col}{row}"].border = Border(bottom=Side(style="thin"))


def double_bottom_border(ws, row):
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws[f"{col}{row}"].border = Border(bottom=Side(style="double"))


def build_excel() -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    # =========================
    # COLUMNS
    # =========================
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 10

    # =========================
    # PRINT HEADER (ALL PAGES)
    # =========================
    ws.oddHeader.left.text = "LOGO 1"
    ws.oddHeader.right.text = "LOGO 2"

    # =========================
    # PAGE 1
    # =========================
    bottom_border(ws, 3)

    ws["A4"] = "Quotation"
    ws["B4"] = datetime.today().strftime("%d.%m.%Y")
    ws["E4"] = "Country"

    ws["A5"] = "Department"
    ws["B5"] = "Sales Person"

    ws["A6"] = "Type"
    ws["B6"] = "X5"

    ws["A7"] = "Protection class"
    ws["B7"] = "VR"

    ws["A8"] = "Top Down"
    ws["B8"] = "Number"

    ws["A9"] = "Model Year"
    ws["B9"] = "YYYY"

    ws["A10"] = "Vehicle Status"
    ws["B10"] = "STOCK / TO ORDER"

    bottom_border(ws, 10)

    ws["D11"] = "Country"
    ws["E11"] = "Page 1"

    ws["B12"] = "Option Code"
    ws["C12"] = "Description"
    ws["E12"] = "Price"

    underline(ws, "B12")
    underline(ws, "C12")

    row = 13

    # ----- BASIC VEHICLE -----
    ws[f"A{row}"] = "Basic Vehicle"
    underline(ws, f"A{row}")
    row += 1

    ws[f"A{row}"] = "Exterior Color"
    underline(ws, f"A{row}")
    row += 1

    ws[f"A{row}"] = "Interior Color"
    underline(ws, f"A{row}")
    row += 1

    ws[f"A{row}"] = "Interior Trim"
    row += 1

    bottom_border(ws, row - 1)

    # ----- STANDARD EQUIPMENT -----
    ws[f"A{row}"] = "Standard Equipment"
    row += 10  # Platzhalter

    # ----- SECURITY EQUIPMENT -----
    ws[f"A{row}"] = "Security Equipment"
    underline(ws, f"A{row}")
    row += 1

    for _ in range(6):
        ws[f"C{row}"] = "Special security line"
        row += 1

    row += 2

    # =========================
    # PAGE 2
    # =========================
    ws.row_breaks.append(Break(id=row))

    ws[f"A{row}"] = "=A8"
    ws[f"B{row}"] = "=B8"
    ws[f"E{row}"] = "Page 2"
    row += 2

    ws[f"A{row}"] = "Optional Equipment"
    underline(ws, f"A{row}")
    row += 10  # Platzhalter

    ws[f"A{row}"] = "Technical Adjustments"
    underline(ws, f"A{row}")
    row += 1

    ws[f"A{row}"] = "Editions"
    underline(ws, f"A{row}")
    row += 2

    ws[f"A{row}"] = "Deletions"
    underline(ws, f"A{row}")
    row += 2

    bottom_border(ws, row)
    row += 1

    ws[f"B{row}"] = "Basic Vehicle Price"
    row += 1

    ws[f"B{row}"] = "Security Package VR6"
    row += 1

    ws[f"B{row}"] = "Optional Equipment"
    row += 1

    ws[f"B{row}"] = "Technical Adjustment"
    row += 0

    bottom_border(ws, row)
    row += 1

    ws[f"B{row}"] = "(Dropdown)"
    row += 1

    ws[f"B{row}"] = "Transportation"
    row += 1

    ws[f"B{row}"] = "Special Discount"
    bottom_border(ws, row)
    row += 1

    ws[f"B{row}"] = "(Dropdown)"
    double_bottom_border(ws, row)
    row += 4

    # =========================
    # PAGE 4 – TECHNICAL DATA
    # =========================
    ws.row_breaks.append(Break(id=row))

    ws[f"A{row}"] = "=A8"
    ws[f"B{row}"] = "=B8"
    ws[f"E{row}"] = "Page 4"
    row += 2

    ws[f"A{row}"] = "Technical Data"
    underline(ws, f"A{row}")
    row += 1

    technical_lines = [
        "Weight",
        "Unladen DIN (without Driver) kg",
        "Unladen EU kg",
        "Gross vehicle weight kg",
        "Engine",
        "Cylinders/valves",
        "Capacity cc3",
        "Output/Engine Speed kW(hp) / rpm",
        "Engine Torque Nm",
        "Performance",
        "Top Speed3 km/h",
        "Acceleration 0-100 km/h s",
        "Fuel Consumption",
        "Combined l/100 km",
        "CO2 emissions g/km"
    ]

    for line in technical_lines:
        ws[f"A{row}"] = line
        ws[f"C{row}"] = "Text / Number"
        row += 1

    # =========================
    # SAVE
    # =========================
    output_dir = os.path.join(os.getcwd(), "output")
    os.makedirs(output_dir, exist_ok=True)

    file_path = os.path.join(output_dir, "BMW_Quotation_V1_2_LAYOUT_ONLY.xlsx")
    wb.save(file_path)

    return file_path
